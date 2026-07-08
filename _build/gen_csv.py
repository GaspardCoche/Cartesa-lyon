#!/usr/bin/env python3.12
"""Exporte l'audit du site en tableur : une ligne par section, colonnes claires
+ colonnes vides « Retour » et « Décision » à remplir. Produit un CSV (toujours)
et un XLSX mis en forme (si openpyxl dispo)."""
import os
import re
import csv
import glob
from html.parser import HTMLParser

NEW = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(NEW, "AUDIT")
SKIP = {"script", "style", "nav", "header", "footer", "svg"}

ORDER = [
    "index.html", "vos-attentes.html", "expertises.html", "coachings.html",
    "formations.html", "ingenierie.html", "rse-decarbonation.html",
    "pourquoi-cartesa.html", "ressourcerie.html", "blog.html", "contact.html",
    "mentions-legales.html", "politique-confidentialite.html", "politique-cookies.html",
]
LABEL = {
    "index.html": "Accueil", "vos-attentes.html": "Vos attentes",
    "expertises.html": "Diagnostic & accompagnements RH", "coachings.html": "Coachings",
    "formations.html": "Formations", "ingenierie.html": "Ingénierie pédagogique",
    "rse-decarbonation.html": "RSE & Décarbonation", "pourquoi-cartesa.html": "Pourquoi CARTESA",
    "ressourcerie.html": "Ressources", "blog.html": "Blog", "contact.html": "Contact",
    "mentions-legales.html": "Mentions légales", "politique-confidentialite.html": "Confidentialité",
    "politique-cookies.html": "Cookies",
}
STATUS = {
    "index.html": "Conservée", "vos-attentes.html": "NOUVELLE",
    "expertises.html": "FUSIONNÉE (Diagnostic+Accompagnements)", "coachings.html": "NOUVELLE",
    "formations.html": "Conservée", "ingenierie.html": "Conservée",
    "rse-decarbonation.html": "Conservée", "pourquoi-cartesa.html": "NOUVELLE (Équipe+Témoignages)",
    "ressourcerie.html": "Conservée", "blog.html": "Conservée", "contact.html": "Conservée",
    "mentions-legales.html": "Légal", "politique-confidentialite.html": "Légal", "politique-cookies.html": "Légal",
}


class Outline(HTMLParser):
    def __init__(self):
        super().__init__()
        self.skip = 0; self.cur = None; self.buf = []; self.textbuf = []
        self.sec_id = None; self.items = []

    def handle_starttag(self, tag, attrs):
        if tag in SKIP: self.skip += 1; return
        if self.skip: return
        a = dict(attrs)
        if tag == "section" and a.get("id"): self.sec_id = a.get("id")
        if tag in ("h1", "h2", "h3"):
            self._flush(); self.cur = tag; self.buf = []

    def handle_endtag(self, tag):
        if tag in SKIP:
            if self.skip: self.skip -= 1
            return
        if self.skip: return
        if tag in ("h1", "h2", "h3") and self.cur == tag:
            text = " ".join("".join(self.buf).split())
            if text: self.items.append({"lvl": int(tag[1]), "text": text, "id": self.sec_id, "content": ""})
            self.cur = None; self.sec_id = None

    def handle_data(self, data):
        if self.skip: return
        (self.buf if self.cur else self.textbuf).append(data)

    def _flush(self):
        if self.items and self.textbuf:
            txt = " ".join("".join(self.textbuf).split())
            if txt: self.items[-1]["content"] = (self.items[-1]["content"] + " " + txt).strip()
        self.textbuf = []


def parse(path):
    html = open(path, encoding="utf-8", errors="ignore").read()
    p = Outline(); p.feed(html); p._flush()
    return p.items


def excerpt(s, n=300):
    s = " ".join(s.split())
    return (s[:n] + "…") if len(s) > n else s


HEADERS = ["N°", "Page", "Fichier", "Statut page", "Niveau", "Section", "Ancre",
           "Contenu (extrait)", "Retour", "Décision (Garder/Modifier/Supprimer/Ajouter)"]

rows = []
n = 0
for page in ORDER:
    path = os.path.join(NEW, page)
    if not os.path.exists(path): continue
    for it in parse(path):
        n += 1
        rows.append([n, LABEL.get(page, page), page, STATUS.get(page, ""),
                     "H" + str(it["lvl"]), it["text"], ("#" + it["id"]) if it.get("id") else "",
                     excerpt(it["content"]), "", ""])

# ── CSV (UTF-8 BOM pour Excel/Numbers, séparateur ; pour Excel FR) ──
csv_path = os.path.join(OUT, "cartesa-audit.csv")
with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
    w = csv.writer(f, delimiter=";")
    w.writerow(HEADERS)
    w.writerows(rows)
print(f"CSV : {csv_path} ({len(rows)} sections)")

# ── XLSX mis en forme (si openpyxl) ──
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "Audit site CARTESA"
    teal = "092933"; gold = "ECD034"
    head_fill = PatternFill("solid", fgColor=teal)
    head_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.append(HEADERS)
    for c in ws[1]:
        c.fill = head_fill; c.font = head_font; c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True); c.border = border
    fills = {"NOUVELLE": "FDF8E8", "FUSIONNÉE (Diagnostic+Accompagnements)": "E4F1F6", "NOUVELLE (Équipe+Témoignages)": "FDF8E8"}
    for r in rows:
        ws.append(r)
        row = ws[ws.max_row]
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True); c.border = border
        # surligne la 1re ligne d'une page (H1) et colore selon statut
        st = r[3]
        if r[4] == "H1":
            for c in row: c.font = Font(bold=True)
        fill = fills.get(st)
        if fill:
            row[3].fill = PatternFill("solid", fgColor=fill)
    widths = [5, 20, 22, 24, 8, 34, 14, 70, 34, 22]
    for i, wdt in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = wdt
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
    xlsx_path = os.path.join(OUT, "cartesa-audit.xlsx")
    wb.save(xlsx_path)
    print(f"XLSX : {xlsx_path}")
except ImportError:
    print("XLSX : openpyxl absent — CSV seul (installe openpyxl pour l'xlsx).")
